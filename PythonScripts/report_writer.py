import json
import os
from collections.abc import Callable
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from config import (
    RESULTS_FOLDER as DATA_FOLDER, CALCULATED_DATA_FILE as INPUT_FILE_NAME, REPORT_FILE as OUTPUT_FILE_NAME,
    ARR_SIZES,
    SORTS_NAMES, ARRAYS_NAMES,
    CALC_TIME_SHEET_NAME, CALC_OP_SHEET_NAME, CHART_SORT_SHEET_NAME, CHART_ARR_SHEET_NAME,
    X_AXIS_TITLE, Y_AXIS_TITLE_TIME,
    CHART_HEIGHT, CHART_WIDTH, CHART_HORIZONTAL_MARGIN, CHART_VERTICAL_MARGIN, Y_AXIS_TITLE_OP,
    FONT_NAME, FONT_SIZE, FONT_COLOR,
)


def read_data() -> dict:
    input_file = DATA_FOLDER + os.sep + INPUT_FILE_NAME
    with open(input_file, 'r') as file:
        data = json.load(file)
    return data


def write_sizes(sheet: Worksheet, column: int = 1):
    sheet.cell(row=1, column=column).value = "Размер"
    sheet.merge_cells(start_row=1, end_row=2, start_column=column, end_column=column)
    row = 3
    for size in ARR_SIZES:
        sheet.cell(row=row, column=column).value = size
        row += 1


def write_arr_names(sheet: Worksheet, start_column: int):
    i = 0
    for name in ARRAYS_NAMES.values():
        sheet.cell(row=2, column=start_column + i).value = name
        i += 1


def get_time(x: dict) -> float:
    return int(x['time_ns']) / 1e6


def get_operations(x: dict) -> float:
    return int(x['operations']) / 1e6


def get_list_of_values(data: dict, get_value: Callable[[dict], int | float]) -> list:
    answer = []
    for arr_size in ARR_SIZES:
        answer.append(get_value(data[str(arr_size)]))
    return answer


def write_values(data: dict, get_value: Callable[[dict], int | float], sheet: Worksheet, column: int):
    start_row = 3
    values = get_list_of_values(data, get_value)
    row = start_row
    for value in values:
        sheet.cell(row=row, column=column).value = value
        row += 1


def write_sort(data: dict, sort_name: str, sheet: Worksheet, get_value: Callable[[dict], int | float], margin: int = 0):
    start_row = 1
    start_column = 1 + margin

    sheet.cell(row=1, column=start_column).value = sort_name
    end_column = start_column + len(ARRAYS_NAMES) - 1
    sheet.merge_cells(start_row=start_row, end_row=start_row, start_column=start_column, end_column=end_column)

    write_arr_names(sheet, start_column)

    column = start_column
    for arr_key in ARRAYS_NAMES.keys():
        write_values(data[arr_key], get_value, sheet, column)
        column += 1


def write_data(data: dict, sheet: Worksheet, get_value: Callable[[dict], int | float]):
    write_sizes(sheet)
    margin = 1
    for sort_id in data.keys():
        write_sort(data[sort_id], SORTS_NAMES[sort_id], sheet, get_value, margin)
        margin += len(ARRAYS_NAMES)


class ChartValues:
    def __init__(self, ref: Reference, title: str = None):
        self.ref = ref
        self.title = title


def draw_line_chart(
        data: list[ChartValues],
        sheet: Worksheet,
        position: str,
        x_axis: Reference = None,
        title: str = None,
        x_axis_title: str = None,
        y_axis_title: str = None
):
    # Create
    chart = LineChart()

    # Fill with data
    for item in data:
        chart.add_data(item.ref)
        if item.title is not None:
            chart.series[-1].title = SeriesLabel(v=item.title)

    if x_axis is not None:
        chart.set_categories(x_axis)

    # Appearance
    if x_axis_title is not None:
        chart.x_axis.title = x_axis_title

    if title is not None:
        chart.title = title

    if y_axis_title is not None:
        chart.y_axis.title = y_axis_title

    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT

    # Save
    sheet.add_chart(chart, position)


def build_sort_chart(
        sheet: Worksheet,
        column: int,
        row: int,
        data_sheet: Worksheet,
        data_start_column: int,
        y_axis_title: str
):
    min_row = 3
    max_row = min_row + len(ARR_SIZES) - 1

    data = [
        ChartValues(
            Reference(worksheet=data_sheet, min_col=data_start_column + i, min_row=min_row, max_row=max_row),
            data_sheet.cell(column=data_start_column + i, row=min_row - 1).value
        ) for i in range(len(ARRAYS_NAMES))
    ]
    position = sheet.cell(column=column, row=row).coordinate
    x_axis = Reference(worksheet=data_sheet, min_col=1, min_row=min_row, max_row=max_row)
    title = data_sheet.cell(column=data_start_column, row=1).value

    draw_line_chart(data, sheet, position, x_axis, title, X_AXIS_TITLE, y_axis_title)


def build_arr_chart(
        sheet: Worksheet,
        column: int,
        row: int,
        data_sheet: Worksheet,
        data_start_column: int,
        y_axis_title: str
):
    min_row = 3
    max_row = min_row + len(ARR_SIZES) - 1

    data = [
        ChartValues(
            Reference(
                worksheet=data_sheet,
                min_col=data_start_column + i * len(ARRAYS_NAMES),
                min_row=min_row,
                max_row=max_row
            ),
            data_sheet.cell(column=2 + i * len(ARRAYS_NAMES), row=min_row - 2).value
        ) for i in range(len(SORTS_NAMES))
    ]
    position = sheet.cell(column=column, row=row).coordinate
    x_axis = Reference(worksheet=data_sheet, min_col=1, min_row=min_row, max_row=max_row)
    title = data_sheet.cell(column=data_start_column, row=2).value

    draw_line_chart(data, sheet, position, x_axis, title, X_AXIS_TITLE, y_axis_title)


def build_sort_chart_column(
        sheet: Worksheet,
        column: int,
        data_sheet: Worksheet,
        y_axis_title: str
):
    for i in range(len(SORTS_NAMES)):
        data_start_column = 2 + i * len(ARRAYS_NAMES)
        row = 1 + i * CHART_VERTICAL_MARGIN
        build_sort_chart(sheet, column, row, data_sheet, data_start_column, y_axis_title)


def build_arr_chart_column(
        sheet: Worksheet,
        column: int,
        data_sheet: Worksheet,
        y_axis_title: str
):
    for i in range(len(ARRAYS_NAMES)):
        data_start_column = 2 + i
        row = 1 + i * CHART_VERTICAL_MARGIN
        build_arr_chart(sheet, column, row, data_sheet, data_start_column, y_axis_title)


def build_sort_charts(wb: Workbook):
    build_sort_chart_column(
        wb[CHART_SORT_SHEET_NAME],
        1,
        wb[CALC_TIME_SHEET_NAME],
        Y_AXIS_TITLE_TIME
    )
    build_sort_chart_column(
        wb[CHART_SORT_SHEET_NAME],
        1 + CHART_HORIZONTAL_MARGIN,
        wb[CALC_OP_SHEET_NAME],
        Y_AXIS_TITLE_OP
    )


def build_arr_charts(wb: Workbook):
    build_arr_chart_column(
        wb[CHART_ARR_SHEET_NAME],
        1,
        wb[CALC_TIME_SHEET_NAME],
        Y_AXIS_TITLE_TIME
    )
    build_arr_chart_column(
        wb[CHART_ARR_SHEET_NAME],
        1 + CHART_HORIZONTAL_MARGIN,
        wb[CALC_OP_SHEET_NAME],
        Y_AXIS_TITLE_OP
    )


def auto_column_width_wo1row(sheet: Worksheet, column: tuple[Cell]):
    max_length = 8
    column_letter = get_column_letter(column[0].column)
    for i in range(1, len(column)):
        cell = column[i]
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    width = max_length * 1.1
    sheet.column_dimensions[column_letter].width = width


def prettier(sheet: Worksheet):
    font = Font(name=FONT_NAME, size=FONT_SIZE, color=FONT_COLOR)
    bold_font = Font(name=FONT_NAME, size=FONT_SIZE, color=FONT_COLOR, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for column in sheet.columns:
        for cell in column:
            cell.font = font
            cell.alignment = alignment
            cell.border = border
        column[0].font = bold_font
        column[1].font = bold_font

    for column in sheet.columns:
        auto_column_width_wo1row(sheet, column)


def write_report(data: dict) -> Workbook:
    wb = Workbook()

    wb.remove(wb.active)
    wb.create_sheet(CALC_TIME_SHEET_NAME, 0)
    wb.create_sheet(CALC_OP_SHEET_NAME, 1)
    wb.create_sheet(CHART_SORT_SHEET_NAME, 2)
    wb.create_sheet(CHART_ARR_SHEET_NAME, 3)

    write_data(data, wb[CALC_TIME_SHEET_NAME], get_time)
    write_data(data, wb[CALC_OP_SHEET_NAME], get_operations)
    prettier(wb[CALC_TIME_SHEET_NAME])
    prettier(wb[CALC_OP_SHEET_NAME])

    build_sort_charts(wb)
    build_arr_charts(wb)

    return wb


def save_report(report: Workbook):
    if not os.path.exists(DATA_FOLDER):
        os.mkdir(DATA_FOLDER)
    output_file = DATA_FOLDER + os.sep + OUTPUT_FILE_NAME
    report.save(output_file)


def main():
    data = read_data()
    report = write_report(data)
    save_report(report)


if __name__ == '__main__':
    main()
